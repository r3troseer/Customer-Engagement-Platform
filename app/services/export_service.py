"""
FR11 — Report export service.
Generates PDF, CSV, and XLSX files from report data dicts.
"""
import csv
import io
from datetime import datetime, timezone


# ── PDF ───────────────────────────────────────────────────────────────────────

def generate_pdf(report_data: dict, report_name: str) -> bytes:
    """
    Build a PDF from report_data using reportlab.
    Header: report title + generation timestamp.
    Body: scalar fields as key/value table; first list field as a data table.
    Returns raw PDF bytes.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm,
                            topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph(report_name.replace("_", " ").title(), styles["Title"]))
    story.append(Paragraph(
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        styles["Normal"],
    ))
    story.append(Spacer(1, 0.5 * cm))

    # Scalar key/value summary table
    scalars = {k: v for k, v in report_data.items() if not isinstance(v, (list, dict))}
    if scalars:
        summary_data = [["Field", "Value"]] + [[str(k), str(v)] for k, v in scalars.items()]
        summary_table = Table(summary_data, colWidths=[7 * cm, 10 * cm])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2d6a4f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f0")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.5 * cm))

    # First list field as full data table
    for key, value in report_data.items():
        if isinstance(value, list) and value and isinstance(value[0], dict):
            story.append(Paragraph(key.replace("_", " ").title(), styles["Heading2"]))
            headers = list(value[0].keys())
            rows = [[str(row.get(h, "")) for h in headers] for row in value]
            detail_data = [headers] + rows
            col_count = len(headers)
            col_width = 17 * cm / col_count
            detail_table = Table(detail_data, colWidths=[col_width] * col_count)
            detail_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#52b788")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f0")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("PADDING", (0, 0), (-1, -1), 5),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]))
            story.append(detail_table)
            break

    doc.build(story)
    return buffer.getvalue()


# ── CSV ───────────────────────────────────────────────────────────────────────

def generate_csv(report_data: dict, report_name: str) -> bytes:
    """
    Build a CSV from report_data.
    Report-type-aware columns for known types; flat key/value fallback.
    Returns UTF-8 bytes.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    rtype = report_name.lower()

    if rtype == "esg":
        objectives = report_data.get("objectives", [])
        writer.writerow(["name", "target", "actual", "unit", "status"])
        for obj in objectives:
            writer.writerow([
                obj.get("name", ""),
                obj.get("target", ""),
                obj.get("actual", ""),
                obj.get("unit", ""),
                obj.get("status", ""),
            ])

    elif rtype == "compliance":
        frameworks = report_data.get("frameworks", [])
        writer.writerow(["framework_name", "compliant", "total", "percentage"])
        for fw in frameworks:
            writer.writerow([
                fw.get("framework_name", ""),
                fw.get("compliant", ""),
                fw.get("total", ""),
                fw.get("percentage", ""),
            ])

    elif rtype == "supplier":
        writer.writerow(["total", "active", "inactive", "suspended",
                         "avg_esg_score", "docs_pending", "certified"])
        writer.writerow([
            report_data.get("total_suppliers", ""),
            report_data.get("active_suppliers", ""),
            report_data.get("inactive_suppliers", ""),
            report_data.get("suspended_suppliers", ""),
            report_data.get("avg_esg_score", ""),
            report_data.get("documents_pending_review", ""),
            report_data.get("certified_suppliers", ""),
        ])

    elif rtype == "dashboard":
        kpis = report_data.get("kpis", [])
        writer.writerow(["kpi_code", "kpi_name", "kpi_value", "unit"])
        for kpi in kpis:
            writer.writerow([
                kpi.get("kpi_code", ""),
                kpi.get("kpi_name", ""),
                kpi.get("kpi_value", ""),
                kpi.get("unit", ""),
            ])

    else:
        # Flat key/value fallback
        writer.writerow(["key", "value"])
        for k, v in report_data.items():
            if not isinstance(v, (list, dict)):
                writer.writerow([k, v])

    return output.getvalue().encode("utf-8")


# ── XLSX ──────────────────────────────────────────────────────────────────────

def generate_xlsx(report_data: dict, report_name: str) -> bytes:
    """
    Build an XLSX workbook using openpyxl.
    Sheet 1 "Summary": scalar key/value pairs.
    Sheet 2 "Details": tabular data per report type.
    Returns raw bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill_dark = PatternFill("solid", fgColor="2D6A4F")
    header_fill_light = PatternFill("solid", fgColor="52B788")

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_summary = wb.active
    assert ws_summary is not None  # Workbook() always creates an active sheet
    ws_summary.title = "Summary"

    ws_summary.append(["Report", report_name.replace("_", " ").title()])
    ws_summary.append(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    ws_summary.append([])
    ws_summary.append(["Field", "Value"])
    for cell in ws_summary[4]:
        cell.font = header_font
        cell.fill = header_fill_dark

    for k, v in report_data.items():
        if not isinstance(v, (list, dict)):
            ws_summary.append([str(k), str(v)])

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 40

    # ── Sheet 2: Details ──────────────────────────────────────────────────────
    ws_detail = wb.create_sheet("Details")
    rtype = report_name.lower()

    if rtype == "esg":
        headers = ["name", "target", "actual", "unit", "status"]
        rows = [[obj.get(h, "") for h in headers] for obj in report_data.get("objectives", [])]
    elif rtype == "compliance":
        headers = ["framework_name", "compliant", "total", "percentage"]
        rows = [[fw.get(h, "") for h in headers] for fw in report_data.get("frameworks", [])]
    elif rtype == "supplier":
        headers = ["total", "active", "inactive", "suspended",
                   "avg_esg_score", "docs_pending", "certified"]
        rows = [[
            report_data.get("total_suppliers", ""),
            report_data.get("active_suppliers", ""),
            report_data.get("inactive_suppliers", ""),
            report_data.get("suspended_suppliers", ""),
            report_data.get("avg_esg_score", ""),
            report_data.get("documents_pending_review", ""),
            report_data.get("certified_suppliers", ""),
        ]]
    elif rtype == "dashboard":
        headers = ["kpi_code", "kpi_name", "kpi_value", "unit"]
        rows = [[kpi.get(h, "") for h in headers] for kpi in report_data.get("kpis", [])]
    else:
        headers = ["key", "value"]
        rows = [[k, v] for k, v in report_data.items() if not isinstance(v, (list, dict))]

    ws_detail.append(headers)
    for cell in ws_detail[1]:
        cell.font = header_font
        cell.fill = header_fill_light
    for row in rows:
        ws_detail.append([str(v) for v in row])

    for col in ws_detail.columns:
        ws_detail.column_dimensions[col[0].column_letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
